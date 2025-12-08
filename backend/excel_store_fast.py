from openpyxl import load_workbook

DEMO_XLSX = "data/Case Study Data_tiny.xlsx"  # change if needed

def get_columns(ws):
    return [cell.value for cell in ws[1]]

def get_next_id_from_sheet(ws, id_col):
    cols = get_columns(ws)
    if id_col not in cols:
        raise Exception(f"{id_col} not found in {ws.title}")

    col_idx = cols.index(id_col) + 1

    # Walk backwards to find last non-empty ID
    for r in range(ws.max_row, 1, -1):
        val = ws.cell(row=r, column=col_idx).value
        if val is not None:
            return int(val) + 1

    return 1

def to_float(val, default=0.0):
    try:
        return float(val)
    except Exception:
        return default

def save_order_from_json(extracted: dict):
    wb = load_workbook(DEMO_XLSX)

    ws_header = wb["SalesOrderHeader"]
    ws_detail = wb["SalesOrderDetail"]

    header_cols = get_columns(ws_header)
    detail_cols = get_columns(ws_detail)

    # 1) Next IDs (fast, no pandas)
    sales_order_id = get_next_id_from_sheet(ws_header, "SalesOrderID")
    detail_id = get_next_id_from_sheet(ws_detail, "SalesOrderDetailID")

    # 2) Line items
    items = extracted.get("lineItems") or []
    for it in items:
        qty = to_float(it.get("qty"), 0)
        unit = to_float(it.get("unitPrice"), 0)
        if it.get("lineTotal") is None:
            it["lineTotal"] = qty * unit

    computed_subtotal = sum(to_float(i.get("lineTotal"), 0) for i in items)
    subtotal = to_float(extracted.get("subtotal"), computed_subtotal)

    tax = extracted.get("tax")
    tax_rate = extracted.get("taxRate")
    if tax is None and tax_rate is not None:
        tax = subtotal * to_float(tax_rate, 0)
    tax = to_float(tax, 0)

    freight = to_float(extracted.get("freight"), 0)

    total_due = extracted.get("totalDue")
    if total_due is None:
        total_due = subtotal + tax + freight
    total_due = to_float(total_due, 0)

    # 3) Build header row dict
    header_row = {
        "SalesOrderID": sales_order_id,
        "RevisionNumber": 1,
        "OrderDate": extracted.get("orderDate"),
        "DueDate": extracted.get("dueDate"),
        "ShipDate": extracted.get("shipDate"),
        "Status": 1,
        "OnlineOrderFlag": False,
        "SalesOrderNumber": f"SO-{sales_order_id}",
        "PurchaseOrderNumber": extracted.get("purchaseOrderNumber") or extracted.get("invoiceNumber"),
        "AccountNumber": (extracted.get("customer") or {}).get("accountNumber"),
        "CustomerID": (extracted.get("customer") or {}).get("customerId") or 0,
        "SubTotal": subtotal,
        "TaxAmt": tax,
        "Freight": freight,
        "TotalDue": total_due,
        "Comment": f"Fast insert: {extracted.get('invoiceNumber', 'N/A')}"
    }

    # 4) Append header in correct column order
    ws_header.append([header_row.get(col, None) for col in header_cols])

    # 5) Append all details (no repeated file open/save)
    for it in items:
        detail_row = {
            "SalesOrderID": sales_order_id,
            "SalesOrderDetailID": detail_id,
            "OrderQty": to_float(it.get("qty"), 0),
            "UnitPrice": to_float(it.get("unitPrice"), 0),
            "UnitPriceDiscount": to_float(it.get("unitPriceDiscount"), 0),
            "LineTotal": to_float(it.get("lineTotal"), 0),
            "ProductID": it.get("productNumber"),
            "CarrierTrackingNumber": None,
            "SpecialOfferID": None
        }
        ws_detail.append([detail_row.get(col, None) for col in detail_cols])
    
    print("DEBUG: saving SalesOrderID =", sales_order_id)
    wb.save(DEMO_XLSX)
    return sales_order_id
