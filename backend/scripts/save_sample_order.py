import pandas as pd
from openpyxl import load_workbook

DEMO_XLSX = "data/Case Study Data_demo.xlsx"

def next_id(sheet_name, id_col):
    df = pd.read_excel(DEMO_XLSX, sheet_name=sheet_name, usecols=[id_col])
    return int(df[id_col].max()) + 1 if not df.empty else 1

def append_row(sheet_name, row_dict):
    wb = load_workbook(DEMO_XLSX)
    ws = wb[sheet_name]

    # column names from the first row
    columns = [cell.value for cell in ws[1]]

    # align values exactly to sheet columns
    row = [row_dict.get(col, None) for col in columns]
    ws.append(row)

    wb.save(DEMO_XLSX)

def save_sample():
    # 1) Generate new IDs
    sales_order_id = next_id("SalesOrderHeader", "SalesOrderID")
    detail_id_1 = next_id("SalesOrderDetail", "SalesOrderDetailID")
    detail_id_2 = detail_id_1 + 1

    # 2) Minimal header (we fill only key financial + date fields)
    header_row = {
        "SalesOrderID": sales_order_id,
        "RevisionNumber": 1,
        "OrderDate": "2014-05-01",
        "DueDate": "2014-05-31",
        "ShipDate": "2014-05-03",
        "Status": 1,
        "OnlineOrderFlag": False,
        "SalesOrderNumber": f"SO-{sales_order_id}",
        "PurchaseOrderNumber": "PO-DEMO-001",
        "AccountNumber": None,
        "CustomerID": 0,
        "SalesPersonID": None,
        "TerritoryID": None,
        "BillToAddressID": None,
        "ShipToAddressID": None,
        "ShipMethodID": None,
        "CreditCardID": None,
        "SubTotal": 200.00,
        "TaxAmt": 10.00,
        "Freight": 0.00,
        "TotalDue": 210.00,
        "Comment": "Demo insert from Python"
    }

    append_row("SalesOrderHeader", header_row)

    # 3) Two sample line items
    detail_row_1 = {
        "SalesOrderDetailID": detail_id_1,
        "SalesOrderID": sales_order_id,
        "CarrierTrackingNumber": None,
        "OrderQty": 1,
        "ProductID": None,
        "SpecialOfferID": None,
        "UnitPrice": 120.00,
        "UnitPriceDiscount": 0.00,
        "LineTotal": 120.00,
    }

    detail_row_2 = {
        "SalesOrderDetailID": detail_id_2,
        "SalesOrderID": sales_order_id,
        "CarrierTrackingNumber": None,
        "OrderQty": 2,
        "ProductID": None,
        "SpecialOfferID": None,
        "UnitPrice": 40.00,
        "UnitPriceDiscount": 0.00,
        "LineTotal": 80.00,
    }

    append_row("SalesOrderDetail", detail_row_1)
    append_row("SalesOrderDetail", detail_row_2)

    print("✅ Saved sample order")
    print("New SalesOrderID:", sales_order_id)
    print("New Detail IDs:", detail_id_1, detail_id_2)

if __name__ == "__main__":
    save_sample()
