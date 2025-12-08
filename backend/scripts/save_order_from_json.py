import json
import pandas as pd
from openpyxl import load_workbook

DEMO_XLSX = "data/Case Study Data_demo.xlsx"
JSON_PATH = "sample_extracted.json"

def next_id(sheet_name, id_col):
    df = pd.read_excel(DEMO_XLSX, sheet_name=sheet_name, usecols=[id_col])
    return int(df[id_col].max()) + 1 if not df.empty else 1

def append_row(sheet_name, row_dict):
    wb = load_workbook(DEMO_XLSX)
    ws = wb[sheet_name]

    columns = [cell.value for cell in ws[1]]
    row = [row_dict.get(col, None) for col in columns]

    ws.append(row)
    wb.save(DEMO_XLSX)

def to_float(val, default=0.0):
    try:
        return float(val)
    except Exception:
        return default

def save_order_from_json(extracted: dict):
    # 1) Create new IDs
    sales_order_id = next_id("SalesOrderHeader", "SalesOrderID")

    # 2) Pull line items safely
    items = extracted.get("lineItems") or []
    for it in items:
        qty = to_float(it.get("qty"), 0)
        unit = to_float(it.get("unitPrice"), 0)
        # If lineTotal missing, compute it
        if it.get("lineTotal") is None:
            it["lineTotal"] = qty * unit

    # 3) Compute totals if missing
    computed_subtotal = sum(to_float(i.get("lineTotal"), 0) for i in items)
    subtotal = to_float(extracted.get("subtotal"), computed_subtotal)

    tax = extracted.get("tax")
    tax_rate = extracted.get("taxRate")
    if tax is None and tax_rate is not None:
        tax = subtotal * to_float(tax_rate, 0)
    tax = to_float(tax, 0)

    freight = to_float(extracted.get("freight"), 0)

    total_due = extracted.get("totalDue")
    if total_due is None:
        total_due = subtotal + tax + freight
    total_due = to_float(total_due, 0)

    # 4) Map Header (your sheet has ~23 columns; we only fill key ones)
    header_row = {
        "SalesOrderID": sales_order_id,
        "RevisionNumber": 1,
        "OrderDate": extracted.get("orderDate"),
        "DueDate": extracted.get("dueDate"),
        "ShipDate": extracted.get("shipDate"),
        "Status": 1,
        "OnlineOrderFlag": False,
        "SalesOrderNumber": f"SO-{sales_order_id}",
        "PurchaseOrderNumber": extracted.get("purchaseOrderNumber") or extracted.get("invoiceNumber"),
        "AccountNumber": extracted.get("customer", {}).get("accountNumber"),
        "CustomerID": extracted.get("customer", {}).get("customerId") or 0,
        "SubTotal": subtotal,
        "TaxAmt": tax,
        "Freight": freight,
        "TotalDue": total_due,
        "Comment": f"Inserted from JSON: {extracted.get('invoiceNumber', 'N/A')}"
    }

    append_row("SalesOrderHeader", header_row)

    # 5) Map Details
    detail_id = next_id("SalesOrderDetail", "SalesOrderDetailID")

    for it in items:
        detail_row = {
            "SalesOrderID": sales_order_id,
            "SalesOrderDetailID": detail_id,
            "OrderQty": to_float(it.get("qty"), 0),
            "UnitPrice": to_float(it.get("unitPrice"), 0),
            "UnitPriceDiscount": to_float(it.get("unitPriceDiscount"), 0),
            "LineTotal": to_float(it.get("lineTotal"), 0),
            # These may exist in your sheet; if not, they'll be ignored
            "ProductID": it.get("productNumber"),
            "CarrierTrackingNumber": None,
            "SpecialOfferID": None
        }
        append_row("SalesOrderDetail", detail_row)
        detail_id += 1

    return sales_order_id

if __name__ == "__main__":
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        extracted = json.load(f)

    new_id = save_order_from_json(extracted)
    print("✅ Saved JSON order to Excel")
    print("New SalesOrderID:", new_id)
